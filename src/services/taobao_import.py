"""千牛 Excel 导入与订单 reconcile 业务逻辑。

模块职责：
1. 解析 .xlsx 文件 → 标准化结构化行（v2 14 列严格表头）
2. 按 14 个字段 / 5 状态映射 / 2 支付方式 / 2 店铺类型分流入账
3. 0.6% 手续费规则：仅 received 时扣（gross - round(gross*0.006, 2)）
4. 微信/received 的 mature_at = ``confirmed_at + 7 天``（精确到分秒，与千牛后台一致）
   —— 4/29 14:30 确认 → 5/6 14:30 进入"聚合可提现"显示
   —— 实际可点提现按钮要等 5/7（与千牛节奏一致，CEO 手动把控,系统不强制拦截）
5. 老订单状态变化 reconcile（撤旧流水 + 建新流水；在途→确认差额=fee 自然消失）
6. 导入末尾自动跑解冻（与原 release_aggregator 端点逻辑一致）
7. 单一事务原子（异常由调用方 rollback）
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from typing import BinaryIO, Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from src.models.taobao import (
    TaobaoOrder,
    TaobaoOrderPaymentMethod,
    TaobaoOrderStatus,
    TaobaoShop,
)
from src.models.wallet import (
    TransactionDirection,
    Wallet,
    WalletTransaction,
    credit,
    debit,
)
from src.services.taobao_maturity import calculate_pending_maturity


# 千牛 v2 导出 Excel 的标准表头（14 列严格匹配，多 1 列、少 1 列、列名错都视为结构错）
EXPECTED_HEADERS: tuple[str, ...] = (
    "订单编号",          # col 0
    "支付单号",          # col 1
    "支付详情",          # col 2
    "买家实付金额",      # col 3
    "订单状态",          # col 4
    "订单创建时间",      # col 5  (忽略,仅占位校验)
    "订单付款时间",      # col 6
    "宝贝种类",          # col 7  (忽略,仅占位校验)
    "店铺名称",          # col 8
    "卖家服务费",        # col 9  (忽略,CEO 同意以淘宝平台 0.2% 计算)
    "退款金额",          # col 10 (忽略,仅占位校验)
    "发货时间",          # col 11
    "确认收货时间",      # col 12
    "确认收货打款金额",  # col 13
)

# 千牛订单状态字面值（_normalize_status 已把全角逗号统一为半角再做比较）
QIANNIU_STATUS_PENDING_PAY = "等待买家付款"
QIANNIU_STATUS_PAID_UNSHIPPED = "买家已付款,等待卖家发货"
QIANNIU_STATUS_SHIPPED_UNCONFIRMED = "卖家已发货,等待买家确认"
QIANNIU_STATUS_RECEIVED = "交易成功"
QIANNIU_STATUS_CLOSED = "交易关闭"

# 微信冻结成熟天数
WECHAT_MATURE_DAYS = 7

# 0.6% 手续费率（淘宝平台抽成，CEO 已确认）
FEE_RATE = Decimal("0.006")
# 2 位精度
TWO_PLACES = Decimal("0.01")


class TaobaoImportError(Exception):
    """文件级错误（结构 / 类型）。路由层翻译为 400/404。"""


@dataclass
class ImportReport:
    """导入报告 —— 路由层 by_alias 序列化为 camelCase。"""

    shop_name: str
    total_rows_parsed: int = 0
    created_orders: int = 0
    status_changed_orders: int = 0
    closed_reverted: int = 0
    skipped_no_change: int = 0
    skipped_unpaid_or_unshipped: int = 0
    skipped_unknown_payment: int = 0
    auto_released_amount: Decimal = field(default_factory=lambda: Decimal("0"))
    auto_released_count: int = 0
    total_fee_amount: Decimal = field(default_factory=lambda: Decimal("0"))
    errors: list[str] = field(default_factory=list)


@dataclass
class ParsedRow:
    """单行解析结果。

    ``system_status``/``payment_method`` 为 None 表示"应跳过"（未付款 / 未知支付）。
    """

    row_index: int  # Excel 行号（1-based,第 1 行表头跳过,数据从 2 起）
    order_number: str
    payment_no: str
    payment_detail_raw: str
    buyer_paid_amount: Decimal
    qianniu_status: str
    paid_at: Optional[datetime]
    shop_name_in_row: str
    shipped_at: Optional[datetime]
    confirmed_at: Optional[datetime]
    confirmed_amount: Decimal
    payment_method: Optional[TaobaoOrderPaymentMethod]
    system_status: Optional[TaobaoOrderStatus]
    is_pending_pay_or_unshipped: bool
    is_unknown_payment: bool


def _to_decimal(value) -> Decimal:
    """空值 / 字符串金额 → Decimal。空字符串当 0。"""
    if value is None or value == "":
        return Decimal("0")
    return Decimal(str(value))


def _to_datetime(value) -> Optional[datetime]:
    """openpyxl 单元格 → datetime。允许空（千牛未发货时 shipped_at 为空）。"""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    # 字符串兜底（一般应被 openpyxl 自动转 datetime,但样例 JSON 是字符串）
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _detect_payment_method(payment_detail: str) -> Optional[TaobaoOrderPaymentMethod]:
    """从 ``支付详情`` 字符串识别支付方式。"""
    if not payment_detail:
        return None
    if "微信支付" in payment_detail:
        return TaobaoOrderPaymentMethod.WECHAT
    if "支付宝" in payment_detail:
        return TaobaoOrderPaymentMethod.ALIPAY
    return None


def _normalize_status(qianniu_status: str) -> str:
    """归一化中文标点（半角/全角逗号统一为半角）后再做字面值比较。"""
    if not qianniu_status:
        return ""
    return qianniu_status.strip().replace("，", ",")


def _map_status(qianniu_status: str) -> tuple[Optional[TaobaoOrderStatus], bool]:
    """千牛状态 → (系统状态, is_pending_or_unshipped)。

    返回 (None, True)  : 等待买家付款 / 买家已付款等待发货 → 跳过且计入 unpaid_or_unshipped
    返回 (status, False): 三个会落地的状态 (shipped_unconfirmed/received/closed)
    返回 (None, False): 完全未知状态（不应发生,作为 errors）
    """
    s = _normalize_status(qianniu_status)
    if s == QIANNIU_STATUS_PENDING_PAY:
        return None, True
    if s == QIANNIU_STATUS_PAID_UNSHIPPED:
        return None, True
    if s == QIANNIU_STATUS_SHIPPED_UNCONFIRMED:
        return TaobaoOrderStatus.SHIPPED_UNCONFIRMED, False
    if s == QIANNIU_STATUS_RECEIVED:
        return TaobaoOrderStatus.RECEIVED, False
    if s == QIANNIU_STATUS_CLOSED:
        return TaobaoOrderStatus.CLOSED, False
    return None, False  # 未知状态


def compute_fee_and_net(gross: Decimal) -> tuple[Decimal, Decimal]:
    """0.2% 手续费规则：fee = round(gross × 0.002, 2) ROUND_HALF_UP；net = gross - fee。

    边界值（CEO 已确认）：
        56.00 → fee=0.11 net=55.89
        269.00 → fee=0.54 net=268.46
        27.50 → fee=0.06 net=27.44 （0.055 → ROUND_HALF_UP → 0.06）
        100.00 → fee=0.20 net=99.80
    """
    g = Decimal(gross)
    raw_fee = g * FEE_RATE
    fee = raw_fee.quantize(TWO_PLACES, rounding=ROUND_HALF_UP)
    net = g - fee
    return fee, net


def parse_workbook(file_obj: BinaryIO) -> list[list]:
    """读取 .xlsx → 第一工作表所有行（含表头）。

    校验：
    1. 必须能被 openpyxl 打开（否则 TaobaoImportError）
    2. 第 1 行（表头）前 14 列必须等于 EXPECTED_HEADERS
    """
    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl 内部多种异常,统一翻译
        raise TaobaoImportError(f"无法解析 .xlsx 文件: {exc}") from exc

    sheet = wb.active
    if sheet is None:
        raise TaobaoImportError("Excel 文件没有可用工作表")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise TaobaoImportError("Excel 文件为空")

    header = rows[0]
    if len(header) < len(EXPECTED_HEADERS):
        raise TaobaoImportError(
            f"Excel 列数不足，期望前 {len(EXPECTED_HEADERS)} 列为 {EXPECTED_HEADERS}"
        )
    actual_first_n = tuple((str(c).strip() if c is not None else "") for c in header[: len(EXPECTED_HEADERS)])
    if actual_first_n != EXPECTED_HEADERS:
        raise TaobaoImportError(
            f"Excel 表头不符，期望 {EXPECTED_HEADERS}，实际 {actual_first_n}"
        )

    return rows


def _parse_row(row_index: int, row: tuple) -> ParsedRow:
    """单行 → ParsedRow（不做 DB 操作）。

    14 列字段映射（仅读取必要列；col 5/7/9/10 占位校验,不读）：
        col 0  订单编号
        col 1  支付单号
        col 2  支付详情
        col 3  买家实付金额
        col 4  订单状态
        col 6  订单付款时间
        col 8  店铺名称
        col 11 发货时间
        col 12 确认收货时间
        col 13 确认收货打款金额
    """
    order_number = str(row[0]).strip() if row[0] is not None else ""
    payment_no = str(row[1]).strip() if row[1] is not None else ""
    payment_detail_raw = str(row[2]) if row[2] is not None else ""
    buyer_paid_amount = _to_decimal(row[3])
    qianniu_status_raw = str(row[4]).strip() if row[4] is not None else ""
    # col 5 订单创建时间 -- 忽略
    paid_at = _to_datetime(row[6])
    # col 7 宝贝种类 -- 忽略
    shop_name_in_row = str(row[8]).strip() if row[8] is not None else ""
    # col 9 卖家服务费 -- 忽略（CEO 同意以平台 0.2% 计算）
    # col 10 退款金额 -- 忽略
    shipped_at = _to_datetime(row[11])
    confirmed_at = _to_datetime(row[12])
    confirmed_amount = _to_decimal(row[13])

    payment_method = _detect_payment_method(payment_detail_raw)
    system_status, is_pending_or_unshipped = _map_status(qianniu_status_raw)
    is_unknown_payment = payment_method is None and not is_pending_or_unshipped and system_status is not None
    # 注意：is_unknown_payment 仅在状态会落地时才追究支付方式

    return ParsedRow(
        row_index=row_index,
        order_number=order_number,
        payment_no=payment_no,
        payment_detail_raw=payment_detail_raw,
        buyer_paid_amount=buyer_paid_amount,
        qianniu_status=qianniu_status_raw,
        paid_at=paid_at,
        shop_name_in_row=shop_name_in_row,
        shipped_at=shipped_at,
        confirmed_at=confirmed_at,
        confirmed_amount=confirmed_amount,
        payment_method=payment_method,
        system_status=system_status,
        is_pending_pay_or_unshipped=is_pending_or_unshipped,
        is_unknown_payment=is_unknown_payment,
    )


def _resolve_target_wallet(
    shop: TaobaoShop,
    payment_method: TaobaoOrderPaymentMethod,
    system_status: TaobaoOrderStatus,
) -> Optional[int]:
    """三维分流（支付方式 × 系统状态）→ 目标钱包 id。

    返回 None 表示该状态不入账（如 closed）。

    分流矩阵（A/B 已统一,不再分流）：
    - alipay/shipped_unconfirmed → unconfirmed_alipay
    - alipay/received            → store_alipay_wallet（A: 资产支付宝子钱包 / B: 兔仔电玩支付宝）
    - wechat/shipped_unconfirmed → unconfirmed_wechat
    - wechat/received            → aggregator_frozen（聚合支付独立于店铺归属）
    - closed                     → None (不入账)
    """
    if system_status == TaobaoOrderStatus.CLOSED:
        return None

    if payment_method == TaobaoOrderPaymentMethod.ALIPAY:
        if system_status == TaobaoOrderStatus.SHIPPED_UNCONFIRMED:
            return shop.unconfirmed_alipay_wallet_id
        if system_status == TaobaoOrderStatus.RECEIVED:
            return shop.store_alipay_wallet_id
    elif payment_method == TaobaoOrderPaymentMethod.WECHAT:
        if system_status == TaobaoOrderStatus.SHIPPED_UNCONFIRMED:
            return shop.unconfirmed_wechat_wallet_id
        if system_status == TaobaoOrderStatus.RECEIVED:
            return shop.aggregator_frozen_wallet_id
    return None


def _gross_for_status(parsed: ParsedRow) -> Decimal:
    """按状态选 gross：received 用确认收货金额；shipped_unconfirmed 用买家实付。"""
    if parsed.system_status == TaobaoOrderStatus.RECEIVED:
        return parsed.confirmed_amount
    if parsed.system_status == TaobaoOrderStatus.SHIPPED_UNCONFIRMED:
        return parsed.buyer_paid_amount
    return Decimal("0")


def _credit_for_row(
    session: Session,
    shop: TaobaoShop,
    parsed: ParsedRow,
    wallet_id: int,
    amount: Decimal,
) -> WalletTransaction:
    """按规则 credit。微信/received 写 mature_at = confirmed_at + 7 天（精确到分秒）。

    ``amount`` 已是入账钱包应入账的"当前金额"（received 是 net；shipped_unconfirmed 是 gross）。

    mature_at 与千牛后台一致：4/29 14:30 确认收货 → 5/6 14:30 进入"聚合可提现"。
    （CEO 沿千牛节奏手动操作提现,系统不强制拦截当天提现）
    """
    mature_at: Optional[datetime] = None
    if (
        parsed.payment_method == TaobaoOrderPaymentMethod.WECHAT
        and parsed.system_status == TaobaoOrderStatus.RECEIVED
    ):
        # 优先用 confirmed_at（Excel "确认收货时间"），缺失时兜底用 shipped_at；都缺则 now()
        base = parsed.confirmed_at or parsed.shipped_at or datetime.now()
        mature_at = base + timedelta(days=WECHAT_MATURE_DAYS)
    remark = f"千牛订单 #{parsed.order_number} {parsed.system_status.value}"
    return credit(
        session,
        wallet_id,
        amount,
        remark=remark,
        mature_at=mature_at,
    )


def _debit_old_tx(
    session: Session,
    order: TaobaoOrder,
    reason: str,
) -> None:
    """根据 ``order.bookkeeping_tx_id`` 找老流水的 amount，从老钱包 debit 同金额。

    若 bookkeeping_tx_id / wallet_id 任一缺失,跳过（视为之前未入账）。
    """
    if order.bookkeeping_tx_id is None or order.bookkeeping_wallet_id is None:
        return
    old_tx = session.get(WalletTransaction, order.bookkeeping_tx_id)
    if old_tx is None:
        return
    debit(
        session,
        order.bookkeeping_wallet_id,
        Decimal(old_tx.amount),
        remark=f"reconcile #{order.order_number} {reason}",
    )


def _auto_release_aggregator(
    session: Session,
    shop: TaobaoShop,
) -> tuple[Decimal, int]:
    """导入流程末尾自动跑解冻。复用 ``calculate_pending_maturity`` 计算 + debit/credit 转账。

    返回 ``(累计金额, 笔数)``。无可解冻流水时返回 ``(Decimal("0"), 0)``,不写流水。

    与原 ``release_aggregator`` 端点的逻辑保持一致：
    - 仅算 mature_at <= now 且仍被某 ``TaobaoOrder.bookkeeping_tx_id`` 引用的流水
    - 累计金额从 frozen debit、credit 到 available
    """
    matured_amount, matured_count = calculate_pending_maturity(
        session, shop.aggregator_frozen_wallet_id
    )
    if matured_count <= 0 or matured_amount <= 0:
        return Decimal("0"), 0

    remark = f"导入自动解冻 {matured_count} 笔到期"
    debit(session, shop.aggregator_frozen_wallet_id, matured_amount, remark=remark)
    credit(session, shop.aggregator_available_wallet_id, matured_amount, remark=remark)
    return matured_amount, matured_count


def import_qianniu_workbook(
    session: Session,
    shop: TaobaoShop,
    file_obj: BinaryIO,
) -> ImportReport:
    """主入口：解析 + 入账 + reconcile + 末尾自动解冻，返回 ImportReport。

    调用方负责 commit / rollback。本函数只 add + flush，**不 commit**。
    任何 SQLAlchemy 异常会向上抛，调用方应 rollback。

    解析时会先把所有数据行解析一次，校验"店铺名称"列与上传 ``shop.name`` 一致，
    任何一行不匹配（含空字符串）→ 抛 TaobaoImportError，路由层翻译为 400 整体回滚。
    """
    rows = parse_workbook(file_obj)
    report = ImportReport(shop_name=shop.name)

    # 第一遍：解析所有数据行 + 校验店铺名称
    parsed_rows: list[tuple[int, ParsedRow]] = []
    parse_errors: list[str] = []
    mismatched_shop_names: set[str] = set()
    for row_index, raw_row in enumerate(rows[1:], start=2):
        try:
            parsed = _parse_row(row_index, raw_row)
        except Exception as exc:
            parse_errors.append(f"行 {row_index} 解析异常: {exc}")
            continue
        parsed_rows.append((row_index, parsed))
        # 行级店铺名校验：空字符串视为不匹配
        if parsed.shop_name_in_row != shop.name:
            mismatched_shop_names.add(parsed.shop_name_in_row)

    if mismatched_shop_names:
        # 排序后展示,稳定输出便于测试
        names = ", ".join(sorted(mismatched_shop_names))
        raise TaobaoImportError(
            f"Excel 包含其他店铺数据：{names}，请上传匹配 {shop.name} 的 Excel"
        )

    # 第二遍：实际入账（解析阶段的解析异常仍写到 report.errors）
    for err in parse_errors:
        report.errors.append(err)
        report.total_rows_parsed += 1

    for row_index, parsed in parsed_rows:
        report.total_rows_parsed += 1

        # 跳过未付款 / 未发货
        if parsed.is_pending_pay_or_unshipped:
            report.skipped_unpaid_or_unshipped += 1
            continue

        # 状态完全未识别（理论不应发生）
        if parsed.system_status is None:
            report.errors.append(
                f"行 {row_index} 订单 {parsed.order_number} 未识别千牛状态: {parsed.qianniu_status!r}"
            )
            continue

        # 支付方式未知（仅当状态会落地时才追究）
        if parsed.is_unknown_payment:
            report.skipped_unknown_payment += 1
            continue

        # 此处 system_status ∈ {shipped_unconfirmed, received, closed}
        existing = session.scalar(
            select(TaobaoOrder).where(TaobaoOrder.order_number == parsed.order_number)
        )

        if existing is None:
            _handle_new_order(session, shop, parsed, report)
        else:
            _handle_existing_order(session, shop, existing, parsed, report)

    # 末尾自动解冻
    auto_amount, auto_count = _auto_release_aggregator(session, shop)
    report.auto_released_amount = auto_amount
    report.auto_released_count = auto_count

    return report


def _handle_new_order(
    session: Session,
    shop: TaobaoShop,
    parsed: ParsedRow,
    report: ImportReport,
) -> None:
    """情况 1：新订单。

    - closed：不入账，但仍记一行 TaobaoOrder（status=closed, bookkeeping=None）
    - received：算 fee/net，credit 用 net；TaobaoOrder.amount=net、gross_amount=gross、fee_amount=fee、confirmed_at
    - shipped_unconfirmed：credit 用 gross；TaobaoOrder.amount=gross、gross_amount=gross、fee_amount=NULL、confirmed_at=NULL
    """
    gross = _gross_for_status(parsed)
    wallet_id = _resolve_target_wallet(shop, parsed.payment_method, parsed.system_status)

    # 默认值（适配 closed / 异常分支）
    fee: Optional[Decimal] = None
    amount = gross

    if parsed.system_status == TaobaoOrderStatus.RECEIVED:
        fee, amount = compute_fee_and_net(gross)

    bookkeeping_wallet_id: Optional[int] = None
    bookkeeping_tx_id: Optional[int] = None

    if parsed.system_status != TaobaoOrderStatus.CLOSED and wallet_id is not None and amount > 0:
        tx = _credit_for_row(session, shop, parsed, wallet_id, amount)
        bookkeeping_wallet_id = wallet_id
        bookkeeping_tx_id = tx.id

    received_at = (
        parsed.confirmed_at or parsed.shipped_at
        if parsed.system_status == TaobaoOrderStatus.RECEIVED
        else None
    )
    confirmed_at = parsed.confirmed_at if parsed.system_status == TaobaoOrderStatus.RECEIVED else None

    order = TaobaoOrder(
        shop_id=shop.id,
        order_number=parsed.order_number,
        payment_method=parsed.payment_method.value if parsed.payment_method is not None else "alipay",
        amount=amount,
        gross_amount=gross,
        fee_amount=fee,
        status=parsed.system_status.value,
        bookkeeping_wallet_id=bookkeeping_wallet_id,
        bookkeeping_tx_id=bookkeeping_tx_id,
        shipped_at=parsed.shipped_at,
        received_at=received_at,
        confirmed_at=confirmed_at,
    )
    session.add(order)
    session.flush()

    if fee is not None:
        report.total_fee_amount += fee
    report.created_orders += 1


def _handle_existing_order(
    session: Session,
    shop: TaobaoShop,
    order: TaobaoOrder,
    parsed: ParsedRow,
    report: ImportReport,
) -> None:
    """情况 2/3/4/5：老订单。

    2: status 没变 → 仅更新 last_synced_at
    3/5: status 变了（包括跳级）→ 撤老流水(按老 amount) + 入新账（received → net；shipped → gross）
    4: 变 closed → 撤老流水 + 不入新账
    """
    old_status_value = order.status.value if hasattr(order.status, "value") else order.status

    if old_status_value == parsed.system_status.value:
        # 情况 2：无变化
        order.last_synced_at = datetime.now()
        report.skipped_no_change += 1
        return

    # 状态发生变化
    if parsed.system_status == TaobaoOrderStatus.CLOSED:
        # 情况 4：变 closed
        _debit_old_tx(session, order, "状态变为关闭")
        order.status = TaobaoOrderStatus.CLOSED.value
        order.bookkeeping_wallet_id = None
        order.bookkeeping_tx_id = None
        order.last_synced_at = datetime.now()
        report.closed_reverted += 1
        return

    # 情况 3/5：变到另一个会落地的状态
    _debit_old_tx(session, order, "状态变化")

    new_gross = _gross_for_status(parsed)
    new_wallet_id = _resolve_target_wallet(shop, parsed.payment_method, parsed.system_status)

    new_fee: Optional[Decimal] = None
    new_amount = new_gross
    if parsed.system_status == TaobaoOrderStatus.RECEIVED:
        new_fee, new_amount = compute_fee_and_net(new_gross)

    new_tx_id: Optional[int] = None
    if new_wallet_id is not None and new_amount > 0:
        new_tx = _credit_for_row(session, shop, parsed, new_wallet_id, new_amount)
        new_tx_id = new_tx.id

    order.status = parsed.system_status.value
    order.amount = new_amount
    order.gross_amount = new_gross
    order.fee_amount = new_fee
    order.bookkeeping_wallet_id = new_wallet_id if new_tx_id is not None else None
    order.bookkeeping_tx_id = new_tx_id
    order.last_synced_at = datetime.now()
    if parsed.system_status == TaobaoOrderStatus.RECEIVED:
        order.received_at = parsed.confirmed_at or parsed.shipped_at or order.received_at
        order.confirmed_at = parsed.confirmed_at or order.confirmed_at
    if parsed.shipped_at is not None:
        order.shipped_at = parsed.shipped_at

    if new_fee is not None:
        report.total_fee_amount += new_fee
    report.status_changed_orders += 1
